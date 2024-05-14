from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, PatternFill, Font, Side, DEFAULT_FONT
from openpyxl.styles.borders import Border

from fuzzywuzzy import fuzz

from modules.status import BcStatus
from modules.fill import ColorCode
from modules.report import OldBomSide, NewBomSide
from modules.errors import MissingRequiredWorksheetError
from modules.ref_des import transform_ref_des
from modules.mapping import transform_to_customer_number
from modules.uom_mapping import get_uom
from modules.utils import *

import os, csv, json


#bc_template_path = os.path.join(os.getcwd(), '_bc_templates')

#bc_template = os.path.join(bc_template_path, 'bc_template.xlsx')
#bc_template = os.path.join(bc_template_path, 'bc_template-compare-shorted bom.xlsx')
#bc_template = os.path.join(bc_template_path, 'bc_template-single_level 2.xlsx')
#bc_template = os.path.join(bc_template_path, 'bc_template - solta_multi level_1.xlsx')
#bc_template = os.path.join(bc_template_path, 'bc_template - IKOS 282080 r4 vs r5 full 2.xlsx')
#bc_template = os.path.join(bc_template_path, 'bc_template - IKOS 281357 r4 vs r7.xlsx')

BASE_DIR = os.path.dirname(__name__)

REPORT_DIR = os.path.join(BASE_DIR, 'bc_reports')

# Template Column Index
# Variables defined below are using only for column access purpose; used as column indies.
# No need for the 'global' keyword as the values will remain unchanged.
LEVEL_COL = 0
PN_COL = 1
DESC_COL = 2
UOM_COL = 3
QTY_COL = 4
REV_COL = 5
REF_DES_COL = 6
MFR_NAME_COL = 7
MFR_PN_COL = 8

# Mfr Clean Name
clean_name_file = os.path.join(os.path.join(BASE_DIR, 'clean_name'), 'agile_clean_name.csv')
CLEAN_NAME_DICT = {}


def load_clean_name(clean_name_file):
    global CLEAN_NAME_DICT
    with open(clean_name_file, 'r') as csv_file:
        csvreader = csv.reader(csv_file)
        _ = next(csvreader)
        for row in csvreader:
            CLEAN_NAME_DICT[row[0]] = row[1]
        csv_file.close()


def load_template(template_file, sheet_name, mapping_setting, bom_type='agile'):
    workbook = load_workbook(template_file, read_only=True)
    sheet = workbook[sheet_name]

    top_level_bom = None
    current_bom = None  # parent bom
    current_row = None
    prev_row = None
    bom_stack = []

    # Column 1 = header
    start_row_index = 2

    # predefine level
    current_level = 0.0
    prev_level = 0.0
    for row in sheet.iter_rows(min_row=start_row_index, max_row=sheet.max_row):
        current_level = row[LEVEL_COL].value

        if current_level is not None:
            parent_number = ''
            # if current_bom is not None:
            #     parent_number = current_bom['part number']
            current_level = float(current_level)
            part_number = str(row[PN_COL].value).strip()
            if bom_type == 'agile':
                part_number = transform_to_customer_number(part_number, mapping_setting)
                
            description = clear_space(row[DESC_COL].value)
            uom = get_uom(row[UOM_COL].value)
            quantity = row[QTY_COL].value
            revision = row[REV_COL].value
            ref_des = row[REF_DES_COL].value
            mfr_name = get_clean_mfr_name(row[MFR_NAME_COL].value)
            mfr_number = row[MFR_PN_COL].value
            
            current_row = parse_to_dict(parent_number, current_level, part_number, description, uom, quantity, revision,
                                ref_des)

            if mfr_name is not None and mfr_number is not None:
                current_row['avl'].append({'mfr name': mfr_name, 'mfr number': mfr_number, 'checked': False ,'compare status': []})

            # Case 1: Handle level 0 - initialize top_level_bom
            if current_level == 0:
                top_level_bom = current_row
                current_bom = top_level_bom
            elif current_level > prev_level:
                if prev_level != 0:
                    bom_stack.append(current_bom) # push current parent onto the stack to switch to the lower level
                current_bom = prev_row  # assigned previous row item to be current parent
                current_row['parent number'] = current_bom['part number']
                current_bom['bom list'].append(current_row)
            elif current_level < prev_level:
                for i in range(int(prev_level - current_level)):
                    current_bom = bom_stack.pop()
                current_row['parent number'] = current_bom['part number']
                current_bom['bom list'].append(current_row)
            elif current_level == prev_level:
                current_row['parent number'] = current_bom['part number']
                current_bom['bom list'].append(current_row)
            prev_level = current_level
            prev_row = current_row
        else:
            mfr_name = get_clean_mfr_name(row[MFR_NAME_COL].value)
            mfr_number = row[MFR_PN_COL].value

            if mfr_name is not None and mfr_number is not None:
                prev_row['avl'].append({'mfr name': mfr_name, 'mfr number': mfr_number, 'checked': False, 'compare status': []})
    workbook.close()
    return top_level_bom

def get_clean_mfr_name(mfr_name):
    if mfr_name is not None:
        if mfr_name in CLEAN_NAME_DICT.keys():
            return clear_space(CLEAN_NAME_DICT[mfr_name])
        else:
            return clear_space(mfr_name)
    return mfr_name

# store BOM data read from bc_template into dictionary
def parse_to_dict(parent_number, level, part_number, description, uom, quantity, revision, ref_des):
    # Handle None/Empty rows
    uom = '' if uom is None else uom
    quantity = '' if quantity is None else float(quantity)
    revision = '' if revision is None else str(revision)

    # create ref des list
    if ref_des is None or ref_des == '':
        ref_des_list = []
    else:
        # transform ref des from string to list
        ref_des_list = transform_ref_des(ref_des)
    
    ref_des = {
        'ref_des list': ref_des_list,
        'add': [],
        'remove': [],
    }
    
    return {
        'parent number': parent_number,
        'level': level,
        'part number': part_number,
        'description': description,
        'uom': uom,
        'quantity': quantity,
        'revision': revision,
        'ref des': ref_des,
        'avl': [],
        'compare status': [],
        'bom list': [],
        'reported flag': False,
    }


def compare_bom(old_bom, new_bom, compare_fields_setting):
    # Compare generic information of Old BOM and New BOM
    compare_item(old_bom, new_bom, compare_fields_setting)
    
    if compare_fields_setting['refdes'] == 1:
        # Compare ref des
        compare_ref_des(old_bom, new_bom)
    
    if compare_fields_setting['mfr'] == 1:
        # Run AVL comparison
        compare_avl(old_bom, new_bom)

    # Compare Bom list of Old BOM to new BOM
    for item_old_bom in old_bom['bom list']:
        parent_number = item_old_bom['parent number']
        level = item_old_bom['level']
        part_number = item_old_bom['part number']
        # item_new_bom = next((item for item in new_bom['bom list'] if
                             # item['part number'] == part_number and item['parent number'] == parent_number and item[
                                 # 'level'] == level), None)
        item_new_bom = None
        for item in new_bom['bom list']:
            if item['parent number'] == parent_number and item['level'] == level and item['part number'] == part_number:
                item_new_bom = item
                break
        
        if item_new_bom is None:
            if BcStatus.PART_REMOVE not in item_old_bom['compare status']:
                item_old_bom['compare status'].append(BcStatus.PART_REMOVE)
                for avl in item_old_bom['avl']:
                    avl['compare status'].append(BcStatus.AVL_REMOVE)
                status_pass_down(item_old_bom, BcStatus.PART_REMOVE, BcStatus.AVL_REMOVE)
            continue
        # compare_item(item_old_bom, item_new_bom)
        else:
            compare_bom(item_old_bom, item_new_bom, compare_fields_setting)
    # Compare Bom list of New BOM to Old BOM
    for item_new_bom in new_bom['bom list']:
        parent_number = item_new_bom['parent number']
        level = item_new_bom['level']
        part_number = item_new_bom['part number']
        # item_old_bom = next((item for item in old_bom['bom list'] if
                             # item['part number'] == part_number and item['parent number'] == parent_number and item[
                                 # 'level'] == level), None)
        item_old_bom = None
        for item in old_bom['bom list']:
            if item['parent number'] == parent_number and item['level'] == level and item['part number'] == part_number:
                item_old_bom = item
                break
        if item_old_bom is None:
            if BcStatus.PART_ADD not in item_new_bom['compare status']:
                item_new_bom['compare status'].append(BcStatus.PART_ADD)
                for avl in item_new_bom['avl']:
                    avl['compare status'].append(BcStatus.AVL_ADD)
                status_pass_down(item_new_bom, BcStatus.PART_ADD, BcStatus.AVL_ADD)
            continue
        else:
            compare_bom(item_old_bom, item_new_bom, compare_fields_setting)


def compare_item(item1, item2, compare_fields_setting):
    if item1['part number'] != item2['part number']:
        if BcStatus.PART_NUMBER_MISMATCH not in item1['compare status']:
            item1['compare status'].append(BcStatus.PART_NUMBER_MISMATCH)
        if BcStatus.PART_NUMBER_MISMATCH not in item2['compare status']:
            item2['compare status'].append(BcStatus.PART_NUMBER_MISMATCH)
    
    if compare_fields_setting['desc'] == 1:
        if item1['description'].lower() != item2['description'].lower():
            if BcStatus.DESCRIPTION_MISMATCH not in item1['compare status']:
                item1['compare status'].append(BcStatus.DESCRIPTION_MISMATCH)
            if BcStatus.DESCRIPTION_MISMATCH not in item2['compare status']:
                item2['compare status'].append(BcStatus.DESCRIPTION_MISMATCH)
    if compare_fields_setting['uom'] == 1:
        # if item1['uom'] != '' and item2['uom'] != '':
        if str(item1['uom']).lower() != str(item2['uom']).lower():
            if BcStatus.UOM_MISMATCH not in item1['compare status']:
                item1['compare status'].append(BcStatus.UOM_MISMATCH)
            if BcStatus.UOM_MISMATCH not in item2['compare status']:
                item2['compare status'].append(BcStatus.UOM_MISMATCH)
    if compare_fields_setting['qty'] == 1:
        if item1['quantity'] != item2['quantity']:
            if BcStatus.QTY_CHANGE not in item1['compare status']:
                item1['compare status'].append(BcStatus.QTY_CHANGE)
            if BcStatus.QTY_CHANGE not in item2['compare status']:
                item2['compare status'].append(BcStatus.QTY_CHANGE)
    if compare_fields_setting['rev'] == 1:
        if item1['revision'] != item2['revision']:
            if BcStatus.REV_CHANGE not in item1['compare status']:
                item1['compare status'].append(BcStatus.REV_CHANGE)
            if BcStatus.REV_CHANGE not in item2['compare status']:
                item2['compare status'].append(BcStatus.REV_CHANGE)


def compare_ref_des(item1, item2):
    item1_ref_des = item1['ref des']
    item1_ref_des_list = item1_ref_des['ref_des list']
    
    item2_ref_des = item2['ref des']
    item2_ref_des_list = item2_ref_des['ref_des list']
    
    # handle empty ref des
    if len(item1_ref_des_list) > 0 and len(item2_ref_des_list) > 0:
        # Handle case ref des is in item 1 but not in item 2 (remove case)
        for ref_des in item1_ref_des_list:
            if ref_des not in item2_ref_des_list:
                if ref_des not in item1_ref_des['remove']:
                    item1_ref_des['remove'].append(ref_des)
                if BcStatus.REF_DES_MISMATCH not in item1['compare status']:
                    item1['compare status'].append(BcStatus.REF_DES_MISMATCH)
        
        for ref_des in item2_ref_des_list:
            if ref_des not in item1_ref_des_list:
                if ref_des not in item2_ref_des['add']:
                    item2_ref_des['add'].append(ref_des)
                if BcStatus.REF_DES_MISMATCH not in item2['compare status']:
                    item2['compare status'].append(BcStatus.REF_DES_MISMATCH)                   
        
    elif len(item1_ref_des_list) > 0 and len(item2_ref_des_list) == 0:
        item1['compare status'].append(BcStatus.REF_DES_MISMATCH)
        item2['compare status'].append(BcStatus.REF_DES_MISMATCH)
    elif len(item1_ref_des_list) == 0 and len(item2_ref_des_list) > 0:
        item2['compare status'].append(BcStatus.REF_DES_MISMATCH)
        item1['compare status'].append(BcStatus.REF_DES_MISMATCH)

def compare_avl(item1, item2):
    TARGET_SCORE = 70
    for avl1 in item1['avl']:
        if avl1['checked'] == False:
            mfr_name_item1 = avl1['mfr name'].lower()
            mfr_number_item1 = avl1['mfr number']
            found_avl_match = False
            mfr_name_match = False
            mfr_number_match = False
            for avl2 in item2['avl']:
                if avl2['checked'] == False:
                    mfr_name_item2 = avl2['mfr name'].lower()
                    match_score = average_match_score(mfr_name_item1, mfr_name_item2)

                    mfr_number_item2 = avl2['mfr number']
                    if match_score >= TARGET_SCORE and str(mfr_number_item1) != str(mfr_number_item2):
                        if BcStatus.MFR_NUMBER_MISMATCH not in avl1['compare status']:
                            avl1['compare status'].append(BcStatus.MFR_NUMBER_MISMATCH)
                        if BcStatus.MFR_NUMBER_MISMATCH not in avl2['compare status']:
                            avl2['compare status'].append(BcStatus.MFR_NUMBER_MISMATCH)
                        remove_status_code(avl1, BcStatus.MFR_NAME_MISMATCH)
                        remove_status_code(avl2, BcStatus.MFR_NAME_MISMATCH)
                        mfr_name_match = True
                    elif match_score < TARGET_SCORE and str(mfr_number_item1) == str(mfr_number_item2):
                        if BcStatus.MFR_NAME_MISMATCH not in avl1['compare status']:
                            avl1['compare status'].append(BcStatus.MFR_NAME_MISMATCH)
                        if BcStatus.MFR_NAME_MISMATCH not in avl2['compare status']:
                            avl2['compare status'].append(BcStatus.MFR_NAME_MISMATCH)
                        remove_status_code(avl1, BcStatus.MFR_NUMBER_MISMATCH)
                        remove_status_code(avl2, BcStatus.MFR_NUMBER_MISMATCH)
                        mfr_number_match = True
                    elif match_score >= TARGET_SCORE and str(mfr_number_item1) == str(mfr_number_item2):
                        avl1['compare status'] = [] # reset the status if they found the match one
                        avl1['checked'] = True
                        avl2['compare status'] = []
                        avl2['checked'] = True
                        found_avl_match = True
                        break
            if not found_avl_match and not mfr_name_match and not mfr_number_match:
                if BcStatus.AVL_REMOVE not in avl1['compare status']:
                    avl1['compare status'].append(BcStatus.AVL_REMOVE)
    
    for avl2 in item2['avl']:
        if avl2['checked'] == False:
            mfr_name_item2 = avl2['mfr name'].lower()
            mfr_number_item2 = avl2['mfr number']
            found_avl_match = False
            mfr_name_match = False
            mfr_number_match = False
            for avl1 in item1['avl']:
                if avl1['checked'] == False:
                    mfr_name_item1 = avl1['mfr name'].lower()
                    match_score = average_match_score(mfr_name_item1, mfr_name_item2)

                    mfr_number_item1 = avl1['mfr number']
                    if match_score >= TARGET_SCORE and str(mfr_number_item2) != str(mfr_number_item1):
                        if BcStatus.MFR_NUMBER_MISMATCH not in avl2['compare status']:
                            avl2['compare status'].append(BcStatus.MFR_NUMBER_MISMATCH)
                        if BcStatus.MFR_NUMBER_MISMATCH not in avl1['compare status']:
                            avl1['compare status'].append(BcStatus.MFR_NUMBER_MISMATCH)
                        remove_status_code(avl1, BcStatus.MFR_NAME_MISMATCH)
                        remove_status_code(avl2, BcStatus.MFR_NAME_MISMATCH)
                        mfr_name_match = True
                    elif match_score < TARGET_SCORE and str(mfr_number_item2) == str(mfr_number_item1):
                        if BcStatus.MFR_NAME_MISMATCH not in avl2['compare status']:
                            avl2['compare status'].append(BcStatus.MFR_NAME_MISMATCH)
                        if BcStatus.MFR_NAME_MISMATCH not in avl1['compare status']:
                            avl1['compare status'].append(BcStatus.MFR_NAME_MISMATCH)
                        remove_status_code(avl1, BcStatus.MFR_NUMBER_MISMATCH)
                        remove_status_code(avl2, BcStatus.MFR_NUMBER_MISMATCH)
                        mfr_number_match = True
                    elif match_score >= TARGET_SCORE and str(mfr_number_item2) == str(mfr_number_item1):
                        avl1['compare status'] = []  # reset the status if they found the match one
                        avl1['checked'] = True
                        avl2['compare status'] = []
                        avl2['checked'] = True
                        found_avl_match = True
                        break
            if not found_avl_match and not mfr_name_match and not mfr_number_match:
                if BcStatus.AVL_ADD not in avl2['compare status']:
                    avl2['compare status'].append(BcStatus.AVL_ADD)

def average_match_score(mfr_name1, mfr_name2):
    simple_ratio = fuzz.ratio(mfr_name1, mfr_name2)
    partial_ratio = fuzz.partial_ratio(mfr_name1, mfr_name2)
    wratio = fuzz.WRatio(mfr_name1, mfr_name2)

    return float((simple_ratio + partial_ratio + wratio)/3)

def remove_status_code(avl, status_code):
    if status_code in avl['compare status']:
        avl['compare status'].remove(status_code)

# This fuction will pass down the ADD/REMOVE code to their sub items
# If a bom is added or removed, that also mean all its sub items are added/removed
def status_pass_down(item, status_code, avl_status_code):
    if item['bom list']:
        for sub_item in item['bom list']:
            sub_item['compare status'].append(status_code)
            if sub_item['avl']:
                for avl in sub_item['avl']:
                    avl['compare status'].append(avl_status_code)
            if sub_item['bom list']:
                status_pass_down(sub_item, status_code, avl_status_code)

# HANDLE COMPARISON REPORT

# Color Schemes
# Old BOM title: F4B084
# Old BOM header: F8CBAD
# New BOM title: 9BC2E6
# New BOM header: BDD7EE
def create_bc_report_template():
    wb = Workbook()
    ws = wb.active
    
    # Create BOM Comparison sheet
    ws.title = 'Report'
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 50
    ws.freeze_panes = "A3"
    report_filename = generate_bc_report_filename()

    center_align = Alignment(horizontal='center')
    title_font = Font(name='Arial', bold=True, size=28)
    old_bom_title_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
    new_bom_title_fill = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
    
    all_around_border = Border(top=Side('thin'), bottom=Side('thin'), left=Side('thin'), right=Side('thin'))
    
    # OLD BOM setting
    ws.merge_cells('A1:I1')
    ws['A1'].value = 'AGILE BOM'
    ws['A1'].alignment = center_align
    ws['A1'].font = title_font
    ws['A1'].fill = old_bom_title_fill
    ws['A1'].border = all_around_border
    create_header(ws, 2, 'A', 'F8CBAD')
    
    # New BOM setting
    ws.merge_cells('K1:S1')
    ws['K1'].value = 'CUSTOMER BOM'
    ws['K1'].alignment = center_align
    ws['K1'].font = title_font
    ws['K1'].fill = new_bom_title_fill
    ws['K1'].border = all_around_border
    create_header(ws, 2, 'K', 'BDD7EE')
    
    ws.column_dimensions['J'].width = 4
    
    # Create Legend Sheet
    all_around_border = Border(top=Side('thin'), bottom=Side('thin'), left=Side('thin'), right=Side('thin'))
    legend_sheet = wb.create_sheet(title='Legend', index=1)
    legend_sheet.sheet_view.showGridLines = False
    legend_sheet.sheet_view.zoomScale = 50
    legend_sheet['B2'].fill = ColorCode.HIGHLIGHT
    legend_sheet['B2'].border = all_around_border
    legend_sheet['C2'].value = 'Parts that have mismatched elements'
    
    legend_sheet['B3'].fill = ColorCode.MISMATCH
    legend_sheet['B3'].border = all_around_border
    legend_sheet['C3'].value = 'Mismatched elements'
    
    legend_sheet['B4'].fill = ColorCode.ADD
    legend_sheet['B4'].border = all_around_border
    legend_sheet['C4'].value = 'Add Part/AVL'
    
    legend_sheet['B5'].fill = ColorCode.REMOVE
    legend_sheet['B5'].border = all_around_border
    legend_sheet['C5'].value = 'Remove Part/AVL'
    
    legend_sheet['B6'].border = all_around_border
    legend_sheet['C6'].value = 'Matched'
    DEFAULT_FONT.size = 18
    wb.save(os.path.join(REPORT_DIR, report_filename))
    return report_filename

def create_header(worksheet, row, col, color):
    headers = ['Level', 'Part Number', 'Description', 'UOM', 'Quantity', 'Revision', 'Ref Des', 'Mfr Name', 'Mfr number']
    top_border = Border(top=Side('thin'))
    header_font = Font(name='Arial', bold=True, size=18)
    header_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    all_around_border = Border(top=Side('thin'), bottom=Side('thin'), left=Side('thin'), right=Side('thin'))
    
    column_size = {'A': 9, 'B': 23, 'C': 23, 'D': 9,'E': 9, 'F': 9, 'G': 9,
                   'H': 30, 'I': 30, 'K': 9, 'L': 23, 'M': 23, 'N': 9, 'O': 9,
                   'P': 9, 'Q': 9, 'R': 30, 'S': 30,}
    
    col_index = 0
    for header in headers:
        column_str = chr(ord(col) + col_index)
        cell = f'{column_str}{row}'
        worksheet[cell].value = header
        worksheet[cell].border = all_around_border
        worksheet[cell].font = header_font
        worksheet[cell].fill = header_fill
        worksheet.column_dimensions[column_str].width = column_size[column_str]
        col_index+=1


def write_report_to_xl(old_bom, new_bom, report_filename):
    report_file = os.path.join(REPORT_DIR, report_filename)
    wb = load_workbook(filename=report_file)
    ws = wb['Report']
    start_row = 3 # start at row 3
    write_item_to_row(ws, start_row, old_bom, new_bom)
    ws.auto_filter.ref = f'A2:S{ws.max_row}'
    wb.save(report_file)


def write_item_to_row(worksheet, row, item1, item2):
    if item1.get('parent number') == item2.get('parent number') and item1.get('level') == item2.get('level') and item1.get('part number') == item2.get('part number'):
        left_row, left_deduct_row = write_item_data(worksheet, row, item1, OldBomSide, NewBomSide)
        right_row, right_reduct_row = write_item_data(worksheet, (left_row - left_deduct_row), item2, NewBomSide, OldBomSide)
        item1['reported flag'] = True
        item2['reported flag'] = True
        row = worksheet.max_row + 1
    elif item1.get('level') is not None:
        row, _ = write_item_data(worksheet, row,  item1, OldBomSide, NewBomSide)
        item1['reported flag'] = True
        row = worksheet.max_row + 1
    elif item2.get('level') is not None:
        row, _ = write_item_data(worksheet, row,  item2, NewBomSide, OldBomSide)
        item2['reported flag'] = True
        row = worksheet.max_row + 1
    sub_bom_list1 = item1.get('bom list', [])
    sub_bom_list2 = item2.get('bom list', [])

    for sub_item1 in sub_bom_list1:
        has_matched_item = False
        for sub_item2 in sub_bom_list2:
            if sub_item1.get('parent number') == sub_item2.get('parent number') and sub_item1.get('level') == sub_item2.get('level') and sub_item1.get('part number') == sub_item2.get('part number'):
                row = write_item_to_row(worksheet, row, sub_item1, sub_item2)
                has_matched_item = True
                break
            # elif not sub_item2['reported flag']:
                # row = write_item_to_row(worksheet, row, {}, sub_item2)
                # break
        if not has_matched_item:
            row = write_item_to_row(worksheet, row, sub_item1, {})

    for sub_item2 in sub_bom_list2:
        has_matched_item = False
        for sub_item1 in sub_bom_list1:
            if sub_item2.get('parent number') == sub_item1.get('parent number') and sub_item2.get(
                'level') == sub_item1.get('level') and sub_item2.get('part number') == sub_item1.get('part number'):
                has_matched_item = True
                break
        if not has_matched_item:
            row = write_item_to_row(worksheet, row, {}, sub_item2)
            # if not sub_item2['reported flag']:
                # row = write_item_to_row(worksheet, row, {}, sub_item2)
    return row
    
    
def write_item_data(worksheet, row, item, bom_side, opposite_bomside):
    level = int(item['level'])
    part_number = item['part number']
    description = item['description']
    uom = item['uom']
    qty = item['quantity']
    rev = item['revision']    
    avl_list = item['avl']
    bc_status = item['compare status']
    
    # handle ref des
    ref_des = item['ref des']
    ref_des_list = ref_des['ref_des list']
    if len(ref_des_list) == 0:
        export_ref_des = ''
    else:
        export_ref_des = ','.join(ref_des_list)    
    
    # Cell styles
    prefer_font = Font(name='Arial', size=18)
    center_align = Alignment(horizontal='center')
    all_around_border = Border(top=Side('thin'), bottom=Side('thin'), left=Side('thin'), right=Side('thin'))
    
    worksheet[f'{bom_side.level}{row}'].value = level
    worksheet[f'{bom_side.level}{row}'].alignment = center_align
    worksheet[f'{bom_side.level}{row}'].font = prefer_font
    worksheet[f'{bom_side.level}{row}'].border = all_around_border
    
    worksheet[f'{bom_side.part_number}{row}'].value = part_number
    worksheet[f'{bom_side.part_number}{row}'].font = prefer_font
    worksheet[f'{bom_side.part_number}{row}'].border = all_around_border
    
    worksheet[f'{bom_side.description}{row}'].value = description
    worksheet[f'{bom_side.description}{row}'].font = prefer_font
    worksheet[f'{bom_side.description}{row}'].border = all_around_border
    
    worksheet[f'{bom_side.uom}{row}'].value = uom
    worksheet[f'{bom_side.uom}{row}'].alignment = center_align
    worksheet[f'{bom_side.uom}{row}'].font = prefer_font
    worksheet[f'{bom_side.uom}{row}'].border = all_around_border
    
    worksheet[f'{bom_side.qty}{row}'].value = qty
    worksheet[f'{bom_side.qty}{row}'].alignment = center_align
    worksheet[f'{bom_side.qty}{row}'].font = prefer_font
    worksheet[f'{bom_side.qty}{row}'].border = all_around_border
    
    worksheet[f'{bom_side.rev}{row}'].value = rev
    worksheet[f'{bom_side.rev}{row}'].alignment = center_align
    worksheet[f'{bom_side.rev}{row}'].font = prefer_font
    worksheet[f'{bom_side.rev}{row}'].border = all_around_border
    
    worksheet[f'{bom_side.ref_des}{row}'].value = export_ref_des
    worksheet[f'{bom_side.ref_des}{row}'].font = prefer_font
    worksheet[f'{bom_side.ref_des}{row}'].border = all_around_border
        
    if len(ref_des['add']) > 0:
        comment_ref_des = f"Add: {','.join(ref_des['add'])}"
        worksheet[f'{bom_side.ref_des}{row}'].comment = Comment(comment_ref_des , author='bom compare')
    if len(ref_des['remove']) > 0:
        comment_ref_des = f"Remove: {','.join(ref_des['remove'])}"
        worksheet[f'{bom_side.ref_des}{row}'].comment = Comment(comment_ref_des , author='bom compare')
    
    # The level will be filled on the opposite side with white color for hidden (Add and Remove only)
    # This will allow to filter out by level on both side
    opposite_font = Font(color='ffffff') 

    # Highlight specific cell.
    for status in bc_status:
        if status == BcStatus.PART_NUMBER_MISMATCH:
            worksheet[f'{bom_side.part_number}{row}'].fill = ColorCode.MISMATCH
            
        elif status == BcStatus.DESCRIPTION_MISMATCH:
            worksheet[f'{bom_side.description}{row}'].fill = ColorCode.MISMATCH
            
        elif status == BcStatus.UOM_MISMATCH:
            worksheet[f'{bom_side.uom}{row}'].fill = ColorCode.MISMATCH
            
        elif status == BcStatus.QTY_CHANGE:
            worksheet[f'{bom_side.qty}{row}'].fill = ColorCode.MISMATCH
            
        elif status == BcStatus.REV_CHANGE:
            worksheet[f'{bom_side.rev}{row}'].fill = ColorCode.MISMATCH
            
        elif status == BcStatus.REF_DES_MISMATCH:
            worksheet[f'{bom_side.ref_des}{row}'].fill = ColorCode.MISMATCH
            
        elif status == BcStatus.PART_ADD:
            # Setup opposite side level for filtering purpose, the level color will be set to white for hidden
            worksheet[f'{opposite_bomside.level}{row}'].value = level
            worksheet[f'{opposite_bomside.level}{row}'].alignment = center_align
            worksheet[f'{opposite_bomside.level}{row}'].font = opposite_font
            #############
            
            worksheet[f'{bom_side.level}{row}'].fill = ColorCode.ADD
            worksheet[f'{bom_side.part_number}{row}'].fill = ColorCode.ADD
            worksheet[f'{bom_side.description}{row}'].fill = ColorCode.ADD
            worksheet[f'{bom_side.uom}{row}'].fill = ColorCode.ADD
            worksheet[f'{bom_side.qty}{row}'].fill = ColorCode.ADD
            worksheet[f'{bom_side.rev}{row}'].fill = ColorCode.ADD
            worksheet[f'{bom_side.ref_des}{row}'].fill = ColorCode.ADD
            worksheet[f'{bom_side.mfr_name}{row}'].fill = ColorCode.ADD
            worksheet[f'{bom_side.mfr_number}{row}'].fill = ColorCode.ADD
        elif status == BcStatus.PART_REMOVE:
            # Setup opposite side level for filtering purpose, the level color will be set to white for hidden
            worksheet[f'{opposite_bomside.level}{row}'].value = level
            worksheet[f'{opposite_bomside.level}{row}'].alignment = center_align
            worksheet[f'{opposite_bomside.level}{row}'].font = opposite_font
            #############
        
            worksheet[f'{bom_side.level}{row}'].fill = ColorCode.REMOVE
            worksheet[f'{bom_side.part_number}{row}'].fill = ColorCode.REMOVE
            worksheet[f'{bom_side.description}{row}'].fill = ColorCode.REMOVE
            worksheet[f'{bom_side.uom}{row}'].fill = ColorCode.REMOVE
            worksheet[f'{bom_side.qty}{row}'].fill = ColorCode.REMOVE
            worksheet[f'{bom_side.rev}{row}'].fill = ColorCode.REMOVE
            worksheet[f'{bom_side.ref_des}{row}'].fill = ColorCode.REMOVE
            worksheet[f'{bom_side.mfr_name}{row}'].fill = ColorCode.REMOVE
            worksheet[f'{bom_side.mfr_number}{row}'].fill = ColorCode.REMOVE

    # handle avl list
    deduct_row = 0
    avl_counter = 0
    for avl in avl_list:
        mfr_name = avl['mfr name']
        mfr_number = avl['mfr number']
        bc_status = avl['compare status']
        worksheet[f'{bom_side.level}{row}'].value = level
        worksheet[f'{bom_side.level}{row}'].alignment = center_align
        worksheet[f'{opposite_bomside.level}{row}'].value = level
        worksheet[f'{opposite_bomside.level}{row}'].alignment = center_align
        if avl_counter > 0: # Set level to white color except for the first AVL in the list
            worksheet[f'{bom_side.level}{row}'].font = opposite_font
            worksheet[f'{opposite_bomside.level}{row}'].font = opposite_font
        
        worksheet[f'{bom_side.mfr_name}{row}'].value = mfr_name
        worksheet[f'{bom_side.mfr_name}{row}'].font = prefer_font
        worksheet[f'{bom_side.mfr_name}{row}'].border = all_around_border
        
        worksheet[f'{bom_side.mfr_number}{row}'].value = mfr_number
        worksheet[f'{bom_side.mfr_number}{row}'].font = prefer_font
        worksheet[f'{bom_side.mfr_number}{row}'].border = all_around_border
        for status in bc_status:
            if status == BcStatus.MFR_NAME_MISMATCH:
                worksheet[f'{bom_side.mfr_name}{row}'].fill = ColorCode.MISMATCH
            elif status == BcStatus.MFR_NUMBER_MISMATCH:
                worksheet[f'{bom_side.mfr_number}{row}'].fill = ColorCode.MISMATCH
            elif status == BcStatus.AVL_ADD:
                worksheet[f'{bom_side.mfr_name}{row}'].fill = ColorCode.ADD
                worksheet[f'{bom_side.mfr_number}{row}'].fill = ColorCode.ADD
            elif status == BcStatus.AVL_REMOVE:
                worksheet[f'{bom_side.mfr_name}{row}'].fill = ColorCode.REMOVE
                worksheet[f'{bom_side.mfr_number}{row}'].fill = ColorCode.REMOVE
        avl_counter += 1
        if len(avl_list) > 1:
            row += 1
            deduct_row += 1
    return (row, deduct_row)

def check_match(item1, item2):
    match = False
    if item1['parent number'] == item2['parent number']:
        if item1['level'] == item2['level']:
            if item1['part number'] == item2['part number']:
                match = True
    return match     

def validate_bc_template(bc_template):
    wb = load_workbook(filename=bc_template, read_only=True)
    has_old_bom_sheet = False
    has_new_bom_sheet = False
    for sheet in wb.worksheets:
        if sheet.title == 'AGILE':
            has_old_bom_sheet = True
            break
    for sheet in wb.worksheets:
        if sheet.title == 'CUSTOMER':
            has_new_bom_sheet = True
            break
    if not has_old_bom_sheet or not has_new_bom_sheet:
        wb.close()
        raise MissingRequiredWorksheetError('Could not find either "Agile BOM" or "Customer BOM" Worksheet in the provided template.')
    wb.close()

def run_bc(bc_template, mapping_setting, compare_fields_setting):
    load_clean_name(clean_name_file)
    agile_bom = load_template(bc_template, 'AGILE', mapping_setting, bom_type='agile')
    cust_bom = load_template(bc_template, 'CUSTOMER', None, bom_type='cust')
    if agile_bom != None and cust_bom != None:
        compare_bom(agile_bom, cust_bom, compare_fields_setting)
        bc_report_file = create_bc_report_template()
        write_report_to_xl(agile_bom, cust_bom, bc_report_file)

        with open("json/agile_bom.json", "w") as outfile:
            json.dump(agile_bom, outfile, indent=4)
        
        with open("json/cust_bom.json", "w") as outfile:
            json.dump(cust_bom, outfile, indent=4)
        os.remove(bc_template)
    return bc_report_file


# if __name__ == '__main__':
    # main()
