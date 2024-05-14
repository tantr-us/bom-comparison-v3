from openpyxl import load_workbook
from datetime import datetime

def generate_bc_report_filename():
    current_time = datetime.now()
    filename = f"{current_time.strftime('%Y-%m-%d_%H%M%S')}_REPORT.xlsx"
    return filename
    
def clear_space(s):
    new_str = ''
    if s == None:
        return ''
    else:
        s = s.strip()
        splitted_str = s.split(' ')
        for word in splitted_str:
            if word != '':
                if len(new_str) == 0:
                    new_str = word
                else:
                    new_str += ' ' + word
        return new_str

def get_header_index_from_xl(bc_template, sheet_name):
    wb = load_workbook(filename=bc_template)
    ws = wb[sheet_name]

    header_index_dict = {}
    for col in ws.iter_cols(min_row=1, max_row=1, max_col=ws.max_column):
        for cell in col:
            header_index_dict[cell.value] = col[0].col_idx-1
    wb.close()
    return header_index_dict
