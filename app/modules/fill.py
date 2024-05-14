from openpyxl.styles import PatternFill

# Bom Comparison Status color code
class ColorCode:
    MISMATCH = PatternFill(start_color='ffdb4e', end_color='ffdb4e', fill_type='solid')
    CHANGE = PatternFill(start_color='d279bb', end_color='d279bb', fill_type='solid')
    ADD = PatternFill(start_color='92d050', end_color='92d050', fill_type='solid')
    REMOVE = PatternFill(start_color='ff6b6f', end_color='ff6b6f', fill_type='solid')
    HIGHLIGHT = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')