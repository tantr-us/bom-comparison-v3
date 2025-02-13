from openpyxl import load_workbook
import json

def get_header_index_from_xl(bc_template, sheet_name):
    wb = load_workbook(filename=bc_template)
    ws = wb[sheet_name]

    header_index_dict = {}
    for col in ws.iter_cols(min_row=1, max_row=1, max_col=ws.max_column):
        for cell in col:
            header_index_dict[cell.value] = col[0].col_idx
    wb.close()
    return header_index_dict

if __name__ == '__main__':
    bc_template = "path_to_BC_template"
    sheet_name = "RAW_BOM_AGILE"
    agile_header_idx_dict = get_header_index_from_xl(bc_template=bc_template, sheet_name=sheet_name)
    cust_header_idx_dict = get_header_index_from_xl(bc_template=bc_template, sheet_name="RAW_BOM_CUSTOMER")
    print(json.dumps(agile_header_idx_dict, indent=4))
    print(json.dumps(cust_header_idx_dict, indent=4))

